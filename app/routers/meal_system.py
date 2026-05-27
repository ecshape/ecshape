"""
API endpoints for the new meal system
Trainers can create meal plans with 3 macros and food options
"""

from collections import defaultdict
from datetime import datetime
from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse, JSONResponse
from sqlalchemy.orm import Session, joinedload
from typing import List, Dict, Any, Optional
import io
import logging

logger = logging.getLogger(__name__)

from app.database import get_db
from app.auth.utils import get_current_user
from app.schemas.auth import UserResponse, UserRole
from app.schemas.meal_system import (
    MealPlanCreate,
    MealPlanUpdate,
    MealPlanResponse,
    CompleteMealPlanCreate,
    MealSlotCreate,
    MealSlotUpdate,
    MealSlotResponse,
    MacroCategoryCreate,
    FoodOptionCreate,
    FoodOptionUpdate,
    FoodOptionResponse,
    ClientMealChoiceCreate,
    ClientMealChoiceUpdate,
    ClientMealChoiceResponse,
    DailyMealHistoryCreate,
    DailyMealHistoryResponse,
    MealHistoryChoiceResponse,
    MealCompletionStatusCreate,
    MealCompletionStatusResponse,
    MealBankCreate,
    MealBankUpdate,
    MealBankResponse,
)
from app.models.user import User
from app.services.trainer_notification_helper import notify_trainer_immediate
from app.services.websocket_service import websocket_service
from app.models.meal_system import (
    MealPlanV2 as NewMealPlan,
    MealSlot,
    MacroCategory,
    FoodOption,
    ClientMealChoice,
    MacroType,
    DailyMealHistory,
    MealBank,
    MealCompletionStatus,
)

router = APIRouter()


def _normalize_date(value: datetime) -> datetime:
    """Normalize a datetime to the start of the day for consistent storage."""
    return value.replace(hour=0, minute=0, second=0, microsecond=0)


# ============ Meal Plan Endpoints ============

@router.post("/plans", response_model=MealPlanResponse, status_code=status.HTTP_201_CREATED)
def create_meal_plan(
    plan_data: MealPlanCreate,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new meal plan (trainer only)"""
    if current_user.role != UserRole.TRAINER and current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only trainers can create meal plans"
        )
    
    # Enforce single active plan per client
    existing_plan = db.query(NewMealPlan).filter(
        NewMealPlan.client_id == plan_data.client_id,
        NewMealPlan.is_active == True
    ).first()

    if existing_plan:
        if existing_plan.trainer_id != current_user.id and current_user.role != UserRole.ADMIN:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Client already has an active meal plan assigned by another trainer"
            )

        updatable_fields = [
            "name",
            "description",
            "number_of_meals",
            "total_calories",
            "protein_target",
            "carb_target",
            "fat_target",
            "start_date",
            "end_date",
        ]

        for field in updatable_fields:
            setattr(existing_plan, field, getattr(plan_data, field))

        existing_plan.trainer_id = current_user.id
        existing_plan.is_active = plan_data.is_active

        db.commit()
        db.refresh(existing_plan)
        return existing_plan

    # Create meal plan
    meal_plan = NewMealPlan(
        client_id=plan_data.client_id,
        trainer_id=current_user.id,
        name=plan_data.name,
        description=plan_data.description,
        number_of_meals=plan_data.number_of_meals,
        total_calories=plan_data.total_calories,
        protein_target=plan_data.protein_target,
        carb_target=plan_data.carb_target,
        fat_target=plan_data.fat_target,
        is_active=plan_data.is_active,
        start_date=plan_data.start_date,
        end_date=plan_data.end_date
    )
    
    db.add(meal_plan)
    db.commit()
    db.refresh(meal_plan)
    
    return meal_plan

@router.post("/plans/complete", response_model=MealPlanResponse, status_code=status.HTTP_201_CREATED)
def create_complete_meal_plan(
    plan_data: CompleteMealPlanCreate,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a complete meal plan with all meals and food options at once (trainer only)"""
    if current_user.role != UserRole.TRAINER and current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only trainers can create meal plans"
        )
    
    existing_plan = db.query(NewMealPlan).filter(
        NewMealPlan.client_id == plan_data.client_id,
        NewMealPlan.is_active == True
    ).first()

    if existing_plan:
        if existing_plan.trainer_id != current_user.id and current_user.role != UserRole.ADMIN:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Client already has an active meal plan assigned by another trainer"
            )

        existing_plan.name = plan_data.name
        existing_plan.description = plan_data.description
        existing_plan.number_of_meals = plan_data.number_of_meals
        existing_plan.total_calories = plan_data.total_calories
        existing_plan.protein_target = plan_data.protein_target
        existing_plan.carb_target = plan_data.carb_target
        existing_plan.fat_target = plan_data.fat_target
        existing_plan.trainer_id = current_user.id
        existing_plan.is_active = True

        existing_plan.meal_slots.clear()
        db.flush()

        target_plan = existing_plan
    else:
        target_plan = NewMealPlan(
            client_id=plan_data.client_id,
            trainer_id=current_user.id,
            name=plan_data.name,
            description=plan_data.description,
            number_of_meals=plan_data.number_of_meals,
            total_calories=plan_data.total_calories,
            protein_target=plan_data.protein_target,
            carb_target=plan_data.carb_target,
            fat_target=plan_data.fat_target
        )
        db.add(target_plan)
        db.flush()

    for order_index, slot_data in enumerate(plan_data.meal_slots):
        meal_slot = MealSlot(
            meal_plan_id=target_plan.id,
            name=slot_data.name,
            order_index=order_index,
            time_suggestion=slot_data.time_suggestion,
            target_calories=slot_data.target_calories,
            target_protein=slot_data.target_protein,
            target_carbs=slot_data.target_carbs,
            target_fat=slot_data.target_fat,
        )
        db.add(meal_slot)
        db.flush()

        for macro_data in slot_data.macro_categories:
            macro_category = MacroCategory(
                meal_slot_id=meal_slot.id,
                macro_type=macro_data.macro_type,
                quantity_instruction=macro_data.quantity_instruction,
                calorie_goal=macro_data.calorie_goal,
                track_cross_macros=macro_data.track_cross_macros if hasattr(macro_data, 'track_cross_macros') else False
            )
            db.add(macro_category)
            db.flush()

            for food_data in macro_data.food_options:
                food_option = FoodOption(
                    macro_category_id=macro_category.id,
                    name=food_data.name,
                    name_hebrew=food_data.name_hebrew,
                    calories=food_data.calories,
                    protein=food_data.protein,
                    carbs=food_data.carbs,
                    fat=food_data.fat,
                    serving_size=food_data.serving_size,
                    measurement_type=food_data.measurement_type if hasattr(food_data, 'measurement_type') else None
                )
                db.add(food_option)

    db.commit()
    db.refresh(target_plan)

    return target_plan

@router.get("/plans", response_model=List[MealPlanResponse])
def get_meal_plans(
    client_id: int = None,
    active_only: bool = True,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get meal plans (trainers see their plans, admins see all, clients see their own)"""
    from app.models.user import User
    
    query = db.query(NewMealPlan)
    
    if current_user.role == UserRole.CLIENT:
        query = query.filter(NewMealPlan.client_id == current_user.id)
    elif current_user.role == UserRole.TRAINER:
        # If trainer queries with client_id, verify the client belongs to them
        if client_id:
            client = db.query(User).filter(User.id == client_id).first()
            if not client or client.trainer_id != current_user.id:
                raise HTTPException(status_code=403, detail="You can only view your clients' meal plans")
        query = query.filter(NewMealPlan.trainer_id == current_user.id)
    # Admins see all
    
    if client_id:
        query = query.filter(NewMealPlan.client_id == client_id)
    
    if active_only:
        query = query.filter(NewMealPlan.is_active == True)
    
    return query.all()

@router.get("/plans/{plan_id}", response_model=MealPlanResponse)
def get_meal_plan(
    plan_id: int,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get a specific meal plan"""
    meal_plan = db.query(NewMealPlan).filter(NewMealPlan.id == plan_id).first()
    
    if not meal_plan:
        raise HTTPException(status_code=404, detail="Meal plan not found")
    
    # Check permissions
    if current_user.role == UserRole.CLIENT and meal_plan.client_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to view this meal plan")
    elif current_user.role == UserRole.TRAINER and meal_plan.trainer_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to view this meal plan")
    
    return meal_plan

@router.put("/plans/{plan_id}", response_model=MealPlanResponse)
def update_meal_plan(
    plan_id: int,
    plan_data: MealPlanUpdate,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update a meal plan (trainer only)"""
    if current_user.role != UserRole.TRAINER and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only trainers can update meal plans")
    
    meal_plan = db.query(NewMealPlan).filter(NewMealPlan.id == plan_id).first()
    
    if not meal_plan:
        raise HTTPException(status_code=404, detail="Meal plan not found")
    
    if current_user.role == UserRole.TRAINER and meal_plan.trainer_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to update this meal plan")
    
    # Update fields
    for field, value in plan_data.dict(exclude_unset=True).items():
        setattr(meal_plan, field, value)
    
    db.commit()
    db.refresh(meal_plan)
    
    return meal_plan

@router.delete("/plans/{plan_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_meal_plan(
    plan_id: int,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete a meal plan (trainer only)"""
    if current_user.role != UserRole.TRAINER and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only trainers can delete meal plans")
    
    meal_plan = db.query(NewMealPlan).filter(NewMealPlan.id == plan_id).first()
    
    if not meal_plan:
        raise HTTPException(status_code=404, detail="Meal plan not found")
    
    if current_user.role == UserRole.TRAINER and meal_plan.trainer_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to delete this meal plan")
    
    db.delete(meal_plan)
    db.commit()
    
    return None

# ============ Meal Slot Endpoints ============

@router.post("/plans/{plan_id}/slots", response_model=MealSlotResponse, status_code=status.HTTP_201_CREATED)
def add_meal_slot(
    plan_id: int,
    slot_data: MealSlotCreate,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Add a meal slot to a plan (trainer only)"""
    if current_user.role != UserRole.TRAINER and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only trainers can add meal slots")
    
    # Verify plan exists and trainer owns it
    meal_plan = db.query(NewMealPlan).filter(NewMealPlan.id == plan_id).first()
    if not meal_plan:
        raise HTTPException(status_code=404, detail="Meal plan not found")
    
    if current_user.role == UserRole.TRAINER and meal_plan.trainer_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    meal_slot = MealSlot(
        meal_plan_id=plan_id,
        name=slot_data.name,
        order_index=slot_data.order_index,
        time_suggestion=slot_data.time_suggestion,
        notes=slot_data.notes,
        target_calories=slot_data.target_calories,
        target_protein=slot_data.target_protein,
        target_carbs=slot_data.target_carbs,
        target_fat=slot_data.target_fat,
    )
    
    db.add(meal_slot)
    db.commit()
    db.refresh(meal_slot)
    
    return meal_slot

# ============ Food Option Endpoints ============

@router.post("/macro-categories/{macro_id}/foods", response_model=FoodOptionResponse, status_code=status.HTTP_201_CREATED)
def add_food_option(
    macro_id: int,
    food_data: FoodOptionCreate,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Add a food option to a macro category (trainer only)"""
    if current_user.role != UserRole.TRAINER and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only trainers can add food options")
    
    from app.models.meal_system import MeasurementType
    food_option = FoodOption(
        macro_category_id=macro_id,
        name=food_data.name,
        name_hebrew=food_data.name_hebrew,
        calories=food_data.calories,
        protein=food_data.protein,
        carbs=food_data.carbs,
        fat=food_data.fat,
        serving_size=food_data.serving_size,
        measurement_type=food_data.measurement_type if hasattr(food_data, 'measurement_type') and food_data.measurement_type else MeasurementType.PER_100G,
        notes=food_data.notes,
        order_index=food_data.order_index
    )
    
    db.add(food_option)
    db.commit()
    db.refresh(food_option)
    
    return food_option

@router.put("/foods/{food_id}", response_model=FoodOptionResponse)
def update_food_option(
    food_id: int,
    food_data: FoodOptionUpdate,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update a food option (trainer only)"""
    if current_user.role != UserRole.TRAINER and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only trainers can update food options")
    
    food_option = db.query(FoodOption).filter(FoodOption.id == food_id).first()
    
    if not food_option:
        raise HTTPException(status_code=404, detail="Food option not found")
    
    for field, value in food_data.dict(exclude_unset=True).items():
        setattr(food_option, field, value)
    
    db.commit()
    db.refresh(food_option)
    
    return food_option

@router.delete("/foods/{food_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_food_option(
    food_id: int,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete a food option (trainer only)"""
    if current_user.role != UserRole.TRAINER and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only trainers can delete food options")
    
    food_option = db.query(FoodOption).filter(FoodOption.id == food_id).first()
    
    if not food_option:
        raise HTTPException(status_code=404, detail="Food option not found")
    
    db.delete(food_option)
    db.commit()
    
    return None

# ============ Client Meal Choice Endpoints (for client tracking) ============

@router.post("/choices", response_model=ClientMealChoiceResponse, status_code=status.HTTP_201_CREATED)
def record_meal_choice(
    choice_data: ClientMealChoiceCreate,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Record a client's meal choice (client only)"""
    if current_user.role != "CLIENT":
        raise HTTPException(status_code=403, detail="Only clients can record meal choices")
    
    # Validate: either food_option_id (for meal plan foods) or custom food fields must be provided
    if not choice_data.food_option_id and not choice_data.custom_food_name:
        raise HTTPException(status_code=400, detail="Either food_option_id or custom_food_name must be provided")
    
    if choice_data.food_option_id and not choice_data.meal_slot_id:
        raise HTTPException(status_code=400, detail="meal_slot_id is required when logging a plan food option")
    
    choice = ClientMealChoice(
        client_id=current_user.id,
        food_option_id=choice_data.food_option_id,
        meal_slot_id=choice_data.meal_slot_id,
        date=choice_data.date,
        quantity=choice_data.quantity,
        photo_path=choice_data.photo_path,
        custom_food_name=choice_data.custom_food_name,
        custom_calories=choice_data.custom_calories,
        custom_protein=choice_data.custom_protein,
        custom_carbs=choice_data.custom_carbs,
        custom_fat=choice_data.custom_fat
    )
    
    db.add(choice)
    db.commit()
    db.refresh(choice)
    
    return choice

@router.get("/choices", response_model=List[ClientMealChoiceResponse])
def get_meal_choices(
    client_id: int = None,
    date: str = None,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get meal choices (trainers see their clients, clients see their own)"""
    from datetime import datetime, timedelta, timezone

    query = db.query(ClientMealChoice)
    
    if current_user.role == "CLIENT":
        query = query.filter(ClientMealChoice.client_id == current_user.id)
    elif current_user.role == UserRole.TRAINER and client_id:
        query = query.filter(ClientMealChoice.client_id == client_id)
    elif client_id:
        query = query.filter(ClientMealChoice.client_id == client_id)
    
    if date:
        try:
            target_date = datetime.fromisoformat(date.replace("Z", "+00:00"))
            if target_date.tzinfo:
                target_date = target_date.astimezone(timezone.utc).replace(tzinfo=None)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format")
        start_of_day = target_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = start_of_day + timedelta(days=1)
        query = query.filter(
            ClientMealChoice.date >= start_of_day,
            ClientMealChoice.date < end_of_day
        )
    
    return query.order_by(ClientMealChoice.date.desc()).all()

@router.put("/choices/{choice_id}", response_model=ClientMealChoiceResponse)
def update_meal_choice(
    choice_id: int,
    choice_data: ClientMealChoiceUpdate,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update a meal choice (trainer for approval, client for their own)"""
    choice = db.query(ClientMealChoice).filter(ClientMealChoice.id == choice_id).first()
    
    if not choice:
        raise HTTPException(status_code=404, detail="Meal choice not found")
    
    # Clients can only update their own choices
    if current_user.role == UserRole.CLIENT and choice.client_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    for field, value in choice_data.dict(exclude_unset=True).items():
        setattr(choice, field, value)
    
    db.commit()
    db.refresh(choice)
    
    return choice

@router.delete("/choices/{choice_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_meal_choice(
    choice_id: int,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete a meal choice"""
    choice = db.query(ClientMealChoice).filter(ClientMealChoice.id == choice_id).first()
    
    if not choice:
        raise HTTPException(status_code=404, detail="Meal choice not found")
    
    if current_user.role == UserRole.CLIENT and choice.client_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    db.delete(choice)
    db.commit()
    
    return None

@router.get("/daily-macros")
def get_daily_macros(
    client_id: int = None,
    date: str = None,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Calculate daily macro consumption for a client on a specific date"""
    from datetime import datetime, timedelta, timezone
    
    # Determine which client to query
    target_client_id = client_id if client_id else current_user.id
    
    # Permission check
    if current_user.role == UserRole.CLIENT and target_client_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Parse date or use today
    if date:
        try:
            parsed_date = datetime.fromisoformat(date.replace('Z', '+00:00'))
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format")
        if parsed_date.tzinfo:
            parsed_date = parsed_date.astimezone(timezone.utc).replace(tzinfo=None)
        target_date = parsed_date
    else:
        target_date = datetime.now()
    
    # Get date range (start and end of day)
    start_of_day = target_date.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_day = start_of_day + timedelta(days=1)
    
    # Get all meal choices for this day
    choices = db.query(ClientMealChoice).filter(
        ClientMealChoice.client_id == target_client_id,
        ClientMealChoice.date >= start_of_day,
        ClientMealChoice.date < end_of_day
    ).all()
    
    # Calculate totals
    total_calories = 0
    total_protein = 0
    total_carbs = 0
    total_fat = 0
    
    for choice in choices:
        # Check if this is a custom food or a meal plan food
        if choice.custom_food_name:
            # Custom food - use stored values directly
            if choice.custom_calories:
                total_calories += choice.custom_calories
            if choice.custom_protein:
                total_protein += choice.custom_protein
            if choice.custom_carbs:
                total_carbs += choice.custom_carbs
            if choice.custom_fat:
                total_fat += choice.custom_fat
        elif choice.food_option_id:
            # Meal plan food - calculate based on quantity
            food_option = db.query(FoodOption).filter(FoodOption.id == choice.food_option_id).first()
            
            if not food_option:
                continue
            
            # Parse quantity (e.g., "150g", "150", "2 slices", "2")
            import re
            from app.models.meal_system import MeasurementType
            
            quantity_value = 1.0  # Default
            if choice.quantity:
                try:
                    # Extract number from string like "150g", "150", "2 slices", "2"
                    match = re.search(r'(\d+(?:\.\d+)?)', choice.quantity)
                    if match:
                        quantity_value = float(match.group(1))
                except:
                    pass
            
            # Calculate macros based on measurement type
            if food_option.measurement_type == MeasurementType.PER_PORTION:
                # For per_portion: multiply nutrition values by number of portions
                scale = quantity_value
            else:
                # For per_100g: scale based on grams consumed (default per 100g)
                grams_consumed = quantity_value
                base_grams = 100
                scale = grams_consumed / base_grams if base_grams > 0 else 1
            
            # Add to totals
            if food_option.calories is not None:
                total_calories += food_option.calories * scale
            if food_option.protein is not None:
                total_protein += food_option.protein * scale
            if food_option.carbs is not None:
                total_carbs += food_option.carbs * scale
            if food_option.fat is not None:
                total_fat += food_option.fat * scale
    
    # Get client's meal plan targets
    meal_plan = db.query(NewMealPlan).filter(
        NewMealPlan.client_id == target_client_id,
        NewMealPlan.is_active == True
    ).first()
    
    targets = {
        "calories": meal_plan.total_calories if meal_plan and meal_plan.total_calories else 2000,
        "protein": meal_plan.protein_target if meal_plan and meal_plan.protein_target else 150,
        "carbs": meal_plan.carb_target if meal_plan and meal_plan.carb_target else 200,
        "fat": meal_plan.fat_target if meal_plan and meal_plan.fat_target else 60
    }
    
    return {
        "date": target_date.date().isoformat(),
        "consumed": {
            "calories": round(total_calories, 1),
            "protein": round(total_protein, 1),
            "carbs": round(total_carbs, 1),
            "fat": round(total_fat, 1)
        },
        "targets": targets,
        "remaining": {
            "calories": round(targets["calories"] - total_calories, 1),
            "protein": round(targets["protein"] - total_protein, 1),
            "carbs": round(targets["carbs"] - total_carbs, 1),
            "fat": round(targets["fat"] - total_fat, 1)
        },
        "percentages": {
            "calories": round((total_calories / targets["calories"] * 100) if targets["calories"] > 0 else 0, 1),
            "protein": round((total_protein / targets["protein"] * 100) if targets["protein"] > 0 else 0, 1),
            "carbs": round((total_carbs / targets["carbs"] * 100) if targets["carbs"] > 0 else 0, 1),
            "fat": round((total_fat / targets["fat"] * 100) if targets["fat"] > 0 else 0, 1)
        }
    }

# ============ Daily Meal History Endpoints ============

@router.post("/history", response_model=DailyMealHistoryResponse, status_code=status.HTTP_201_CREATED)
def save_daily_meal_history(
    history_data: DailyMealHistoryCreate,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Save or update daily meal history"""
    from datetime import datetime
    
    # Ensure clients can only save their own history
    client_id = history_data.client_id if history_data.client_id else current_user.id
    if current_user.role == UserRole.CLIENT and client_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if history already exists for this date
    start_of_day = history_data.date.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_day = start_of_day.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    existing_history = db.query(DailyMealHistory).filter(
        DailyMealHistory.client_id == client_id,
        DailyMealHistory.date >= start_of_day,
        DailyMealHistory.date <= end_of_day
    ).first()
    
    if existing_history:
        # Update existing
        existing_history.total_calories = history_data.total_calories
        existing_history.total_protein = history_data.total_protein
        existing_history.total_carbs = history_data.total_carbs
        existing_history.total_fat = history_data.total_fat
        existing_history.is_complete = history_data.is_complete
        db.commit()
        db.refresh(existing_history)
        return existing_history
    else:
        # Create new
        new_history = DailyMealHistory(
            client_id=client_id,
            date=history_data.date,
            total_calories=history_data.total_calories,
            total_protein=history_data.total_protein,
            total_carbs=history_data.total_carbs,
            total_fat=history_data.total_fat,
            is_complete=history_data.is_complete
        )
        db.add(new_history)
        db.commit()
        db.refresh(new_history)
        return new_history

@router.get("/history", response_model=List[DailyMealHistoryResponse])
def get_meal_history(
    client_id: int = None,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get meal history for a client"""
    target_client_id = client_id if client_id else current_user.id
    
    # Permission check
    if current_user.role == UserRole.CLIENT and target_client_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    history_entries = db.query(DailyMealHistory).filter(
        DailyMealHistory.client_id == target_client_id
    ).order_by(DailyMealHistory.date.desc()).all()
    
    from datetime import timedelta
    import re

    detailed_history: List[Dict[str, Any]] = []

    for entry in history_entries:
        start_of_day = entry.date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = start_of_day + timedelta(days=1)

        choices = (
            db.query(ClientMealChoice)
            .options(
                joinedload(ClientMealChoice.food_option)
                .joinedload(FoodOption.macro_category)
                .joinedload(MacroCategory.meal_slot)
            )
            .filter(
                ClientMealChoice.client_id == target_client_id,
                ClientMealChoice.date >= start_of_day,
                ClientMealChoice.date < end_of_day,
            )
            .all()
        )

        slot_ids = {choice.meal_slot_id for choice in choices if choice.meal_slot_id}
        slot_map: Dict[int, MealSlot] = {}
        if slot_ids:
            slot_map = {
                slot.id: slot
                for slot in db.query(MealSlot).filter(MealSlot.id.in_(slot_ids)).all()
            }

        completion_rows = (
            db.query(MealCompletionStatus)
            .filter(
                MealCompletionStatus.client_id == target_client_id,
                MealCompletionStatus.date == _normalize_date(start_of_day),
            )
            .all()
        )
        completion_map = {row.meal_slot_id: row.is_completed for row in completion_rows}

        meals_map: Dict[str, Dict[str, Any]] = {}

        def parse_quantity(quantity: Optional[str]) -> float:
            if not quantity:
                return 100.0
            match = re.search(r"(\d+(?:\.\d+)?)", quantity)
            if match:
                try:
                    return float(match.group(1))
                except ValueError:
                    return 100.0
            return 100.0

        def scale_value(value: Optional[float], scale: float) -> Optional[float]:
            if value is None:
                return None
            return round(value * scale, 1)

        for choice in choices:
            associated_slot = None
            if choice.meal_slot_id:
                associated_slot = slot_map.get(choice.meal_slot_id)
            if not associated_slot and choice.food_option:
                macro_category = choice.food_option.macro_category
                if macro_category and macro_category.meal_slot:
                    associated_slot = macro_category.meal_slot

            if associated_slot:
                meal_key = f"slot-{associated_slot.id}"
                meal_entry = meals_map.setdefault(
                    meal_key,
                    {
                        "meal_slot_id": associated_slot.id,
                        "meal_name": associated_slot.name,
                        "time_suggestion": associated_slot.time_suggestion,
                        "choices": [],
                        "order_index": associated_slot.order_index or 0,
                        "is_completed": completion_map.get(associated_slot.id, False),
                    },
                )
            else:
                meal_entry = meals_map.setdefault(
                    "custom",
                    {
                        "meal_slot_id": None,
                        "meal_name": "Custom Entries",
                        "time_suggestion": None,
                        "choices": [],
                        "order_index": 999,
                        "is_completed": False,
                    },
                )

            calories_value: Optional[float] = None
            protein_value: Optional[float] = None
            carbs_value: Optional[float] = None
            fat_value: Optional[float] = None
            macro_type_value: Optional[MacroType] = None
            food_name: Optional[str] = None
            food_name_hebrew: Optional[str] = None

            if choice.food_option:
                food_option = choice.food_option
                food_name = food_option.name
                food_name_hebrew = food_option.name_hebrew
                macro_category = food_option.macro_category
                if macro_category:
                    macro_type_value = macro_category.macro_type

                from app.models.meal_system import MeasurementType
                import re
                
                # Parse quantity
                quantity_value = parse_quantity(choice.quantity)
                
                # Calculate scale based on measurement type
                if food_option.measurement_type == MeasurementType.PER_PORTION:
                    # For per_portion: multiply by number of portions
                    scale = quantity_value
                else:
                    # For per_100g: scale based on grams (default per 100g)
                    scale = quantity_value / 100 if quantity_value else 1

                calories_value = scale_value(food_option.calories, scale)
                protein_value = scale_value(food_option.protein, scale)
                carbs_value = scale_value(food_option.carbs, scale)
                fat_value = scale_value(food_option.fat, scale)
            else:
                food_name = choice.custom_food_name
                calories_value = (
                    round(choice.custom_calories, 1) if choice.custom_calories is not None else None
                )
                protein_value = (
                    round(choice.custom_protein, 1) if choice.custom_protein is not None else None
                )
                carbs_value = (
                    round(choice.custom_carbs, 1) if choice.custom_carbs is not None else None
                )
                fat_value = (
                    round(choice.custom_fat, 1) if choice.custom_fat is not None else None
                )

            choice_payload = MealHistoryChoiceResponse(
                choice_id=choice.id,
                food_option_id=choice.food_option_id,
                meal_slot_id=associated_slot.id if associated_slot else None,
                macro_type=macro_type_value,
                food_name=food_name,
                food_name_hebrew=food_name_hebrew,
                quantity=choice.quantity,
                is_custom=bool(choice.custom_food_name),
                calories=calories_value,
                protein=protein_value,
                carbs=carbs_value,
                fat=fat_value,
                is_approved=choice.is_approved,
                trainer_comment=choice.trainer_comment,
                photo_path=choice.photo_path,
            )

            meal_entry["choices"].append(choice_payload.dict())

        meals_list: List[Dict[str, Any]] = sorted(
            meals_map.values(), key=lambda meal: (meal["order_index"], meal["meal_name"])
        )

        for meal in meals_list:
            meal.pop("order_index", None)

        detailed_history.append(
            {
                "id": entry.id,
                "client_id": entry.client_id,
                "date": entry.date,
                "total_calories": entry.total_calories,
                "total_protein": entry.total_protein,
                "total_carbs": entry.total_carbs,
                "total_fat": entry.total_fat,
                "is_complete": entry.is_complete,
                "created_at": entry.created_at,
                "updated_at": entry.updated_at,
                "meals": meals_list,
            }
        )

    return detailed_history

@router.get("/history/average")
def get_average_calories(
    days: int = 7,
    client_id: int = None,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Calculate average calories per day over specified period"""
    from datetime import datetime, timedelta
    from sqlalchemy import func
    
    target_client_id = client_id if client_id else current_user.id
    
    # Permission check
    if current_user.role == UserRole.CLIENT and target_client_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Calculate date range
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)
    
    # Get daily histories in date range
    histories = db.query(DailyMealHistory).filter(
        DailyMealHistory.client_id == target_client_id,
        DailyMealHistory.date >= start_date,
        DailyMealHistory.date <= end_date
    ).all()
    
    if not histories:
        return {
            "average_calories": 0,
            "total_days": 0,
            "period": f"Last {days} days"
        }
    
    total_calories = sum(h.total_calories for h in histories)
    average_calories = total_calories / len(histories) if len(histories) > 0 else 0
    
    return {
        "average_calories": round(average_calories, 1),
        "total_calories": round(total_calories, 1),  # Sum of all calories in the period
        "total_days": len(histories),
        "period": f"Last {days} days",
        "detail_history": [
            {
                "date": h.date.isoformat(),
                "calories": h.total_calories,
                "is_complete": h.is_complete
            } for h in histories
        ]
    }

# ============ Meal Completion Endpoints ============


@router.get("/completions", response_model=List[MealCompletionStatusResponse])
def get_meal_completions(
    date: Optional[str] = None,
    client_id: Optional[int] = None,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Fetch completion state for all meals on a specific date.
    Clients see their own data. Trainers/Admins can query their clients.
    """
    target_client_id = client_id if client_id is not None else current_user.id

    if current_user.role == UserRole.CLIENT and target_client_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")

    if current_user.role == UserRole.TRAINER:
        plan_exists = db.query(NewMealPlan).filter(
            NewMealPlan.client_id == target_client_id,
            NewMealPlan.trainer_id == current_user.id,
            NewMealPlan.is_active == True,
        ).first()
        if not plan_exists:
            raise HTTPException(status_code=403, detail="Not authorized")

    target_date = datetime.utcnow()
    if date:
        try:
            target_date = datetime.fromisoformat(date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format")

    normalized_date = _normalize_date(target_date)

    statuses = (
        db.query(MealCompletionStatus)
        .filter(
            MealCompletionStatus.client_id == target_client_id,
            MealCompletionStatus.date == normalized_date,
        )
        .all()
    )

    return statuses


@router.post("/completions", response_model=MealCompletionStatusResponse)
def upsert_meal_completion(
    completion_data: MealCompletionStatusCreate,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Create or update a completion state for a specific meal slot/date.
    """
    target_client_id = completion_data.client_id or current_user.id

    if current_user.role == UserRole.CLIENT and target_client_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")

    slot = (
        db.query(MealSlot)
        .options(joinedload(MealSlot.meal_plan))
        .filter(MealSlot.id == completion_data.meal_slot_id)
        .first()
    )

    if not slot or not slot.meal_plan:
        raise HTTPException(status_code=404, detail="Meal slot not found")

    if slot.meal_plan.client_id != target_client_id:
        raise HTTPException(status_code=400, detail="Meal slot does not belong to target client")

    if current_user.role == UserRole.TRAINER and slot.meal_plan.trainer_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")

    normalized_date = _normalize_date(completion_data.date)

    existing_status = (
        db.query(MealCompletionStatus)
        .filter(
            MealCompletionStatus.client_id == target_client_id,
            MealCompletionStatus.meal_slot_id == completion_data.meal_slot_id,
            MealCompletionStatus.date == normalized_date,
        )
        .first()
    )

    completed_at = datetime.utcnow() if completion_data.is_completed else None
    background_tasks = BackgroundTasks()

    if existing_status:
        existing_status.is_completed = completion_data.is_completed
        existing_status.completion_method = completion_data.completion_method
        existing_status.completed_at = completed_at
        db.commit()
        db.refresh(existing_status)
        if completion_data.is_completed:
            client = db.query(User).filter(User.id == target_client_id).first()
            client_name = (client.full_name or client.username) if client else "Client"
            created = notify_trainer_immediate(
                db,
                target_client_id,
                title="Meal completed",
                message=f"{client_name} completed a meal",
                event_type="meal_completion",
                notification_type="success",
            )
            if created:
                background_tasks.add_task(websocket_service.send_new_notification_hint, created.recipient_id)
        body = MealCompletionStatusResponse.model_validate(existing_status)
        return JSONResponse(content=body.model_dump(mode="json"), background=background_tasks)

    new_status = MealCompletionStatus(
        client_id=target_client_id,
        meal_slot_id=completion_data.meal_slot_id,
        date=normalized_date,
        is_completed=completion_data.is_completed,
        completion_method=completion_data.completion_method,
        completed_at=completed_at,
    )
    db.add(new_status)
    db.commit()
    db.refresh(new_status)
    if completion_data.is_completed:
        client = db.query(User).filter(User.id == target_client_id).first()
        client_name = (client.full_name or client.username) if client else "Client"
        created = notify_trainer_immediate(
            db,
            target_client_id,
            title="Meal completed",
            message=f"{client_name} completed a meal",
            event_type="meal_completion",
            notification_type="success",
        )
        if created:
            background_tasks.add_task(websocket_service.send_new_notification_hint, created.recipient_id)
    body = MealCompletionStatusResponse.model_validate(new_status)
    return JSONResponse(content=body.model_dump(mode="json"), background=background_tasks)

# ============ Meal Bank Endpoints ============

@router.post("/meal-bank", response_model=MealBankResponse, status_code=status.HTTP_201_CREATED)
def create_meal_bank_item(
    item_data: MealBankCreate,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new meal bank item (trainer only)"""
    if current_user.role != UserRole.TRAINER and current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only trainers can create meal bank items"
        )
    
    english_name = item_data.name.strip()
    hebrew_name = item_data.name_hebrew.strip() if item_data.name_hebrew else ""

    if not english_name and not hebrew_name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="At least one of name or name_hebrew must be provided"
        )

    from app.models.meal_system import MeasurementType
    meal_bank_item = MealBank(
        name=english_name or hebrew_name,
        name_hebrew=hebrew_name or None,
        macro_type=item_data.macro_type,
        calories=item_data.calories,
        protein=item_data.protein,
        carbs=item_data.carbs,
        fat=item_data.fat,
        measurement_type=item_data.measurement_type if hasattr(item_data, 'measurement_type') and item_data.measurement_type else MeasurementType.PER_100G,
        serving_size=item_data.serving_size if hasattr(item_data, 'serving_size') else None,
        created_by=current_user.id,
        is_public=item_data.is_public
    )
    
    db.add(meal_bank_item)
    db.commit()
    db.refresh(meal_bank_item)
    
    return meal_bank_item

@router.get("/meal-bank", response_model=List[MealBankResponse])
def get_meal_bank_items(
    trainer_id: int = None,
    macro_type: MacroType = None,
    search: str = None,
    include_public: bool = True,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get meal bank items (trainers and admins see all items, clients see only public)"""
    try:
        query = db.query(MealBank)
        
        if current_user.role == UserRole.TRAINER or current_user.role == UserRole.ADMIN:
            # Trainers and admins see all items
            if not include_public:
                # If include_public is False, only show items created by current user
                query = query.filter(MealBank.created_by == current_user.id)
        else:
            # Clients only see public items
            query = query.filter(MealBank.is_public == True)
        
        # Only apply trainer_id filter if explicitly provided
        if trainer_id:
            query = query.filter(MealBank.created_by == trainer_id)
        
        if macro_type:
            query = query.filter(MealBank.macro_type == macro_type)
        
        if search:
            from sqlalchemy import or_
            query = query.filter(
                or_(
                    MealBank.name.contains(search),
                    MealBank.name_hebrew.contains(search)
                )
            )
        
        items = query.order_by(MealBank.name).all()
        
        # Ensure all items have measurement_type set (for backward compatibility)
        from app.models.meal_system import MeasurementType
        for item in items:
            if not hasattr(item, 'measurement_type') or item.measurement_type is None:
                item.measurement_type = MeasurementType.PER_100G
        
        return items
    except Exception as e:
        logger.error(f"Error fetching meal bank items: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch meal bank items: {str(e)}"
        )

@router.get("/meal-bank/{item_id}", response_model=MealBankResponse)
def get_meal_bank_item(
    item_id: int,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get a specific meal bank item"""
    meal_bank_item = db.query(MealBank).filter(MealBank.id == item_id).first()
    
    if not meal_bank_item:
        raise HTTPException(status_code=404, detail="Meal bank item not found")
    
    # Check if user has permission to view
    if meal_bank_item.created_by != current_user.id and not meal_bank_item.is_public:
        if current_user.role != UserRole.ADMIN:
            raise HTTPException(status_code=403, detail="Not authorized to view this item")
    
    return meal_bank_item

@router.put("/meal-bank/{item_id}", response_model=MealBankResponse)
def update_meal_bank_item(
    item_id: int,
    item_data: MealBankUpdate,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update a meal bank item (trainer only, own items)"""
    if current_user.role != UserRole.TRAINER and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only trainers can update meal bank items")
    
    meal_bank_item = db.query(MealBank).filter(MealBank.id == item_id).first()
    
    if not meal_bank_item:
        raise HTTPException(status_code=404, detail="Meal bank item not found")
    
    if current_user.role == UserRole.TRAINER and meal_bank_item.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to update this item")
    
    # Update fields
    update_data = item_data.dict(exclude_unset=True)

    if "name" in update_data or "name_hebrew" in update_data:
        english_name = update_data.get("name", meal_bank_item.name or "")
        hebrew_name = update_data.get("name_hebrew", meal_bank_item.name_hebrew or "")

        english_name = english_name.strip() if english_name else ""
        hebrew_name = hebrew_name.strip() if hebrew_name else ""

        if not english_name and not hebrew_name:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="At least one of name or name_hebrew must be provided"
            )

        meal_bank_item.name = english_name or hebrew_name
        meal_bank_item.name_hebrew = hebrew_name or None

        # Remove processed fields so they are not set again
        update_data.pop("name", None)
        update_data.pop("name_hebrew", None)

    for field, value in update_data.items():
        setattr(meal_bank_item, field, value)
    
    db.commit()
    db.refresh(meal_bank_item)
    
    return meal_bank_item

@router.delete("/meal-bank/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_meal_bank_item(
    item_id: int,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete a meal bank item (trainer only, own items)"""
    if current_user.role != UserRole.TRAINER and current_user.role != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Only trainers can delete meal bank items")
    
    meal_bank_item = db.query(MealBank).filter(MealBank.id == item_id).first()
    
    if not meal_bank_item:
        raise HTTPException(status_code=404, detail="Meal bank item not found")
    
    if current_user.role == UserRole.TRAINER and meal_bank_item.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to delete this item")
    
    db.delete(meal_bank_item)
    db.commit()
    
    return None

@router.post("/meal-bank/cleanup-duplicates")
def cleanup_duplicate_meal_bank_items(
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Find and remove duplicate meal bank items (trainer/admin only)"""
    if current_user.role != UserRole.TRAINER and current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only trainers and admins can cleanup duplicates"
        )
    
    try:
        import unicodedata
        import re
        
        # Hebrew text normalization function
        def normalize_hebrew(text: str) -> str:
            """Normalize Hebrew text for matching (remove diacritics, handle variations)"""
            if not text:
                return ""
            text = unicodedata.normalize('NFKD', text)
            text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
            text = re.sub(r'\s+', ' ', text.strip().lower())
            return text
        
        # Get all meal bank items
        all_items = db.query(MealBank).all()
        
        # Group items by normalized name
        seen = {}
        duplicates_to_remove = []
        
        for item in all_items:
            # Normalize both Hebrew and English names
            norm_hebrew = normalize_hebrew(item.name_hebrew) if item.name_hebrew else ""
            norm_name = normalize_hebrew(item.name) if item.name else ""
            
            # Use Hebrew name as primary key if available, otherwise English
            key = norm_hebrew if norm_hebrew else norm_name
            
            if not key:
                continue
            
            if key in seen:
                # Found a duplicate - keep the one with more complete data or older one
                existing = seen[key]
                
                # Prefer item with more complete data (has both names, more macros, etc.)
                existing_score = (
                    (1 if existing.name_hebrew else 0) +
                    (1 if existing.name else 0) +
                    (1 if existing.calories else 0) +
                    (1 if existing.protein else 0) +
                    (1 if existing.carbs else 0) +
                    (1 if existing.fat else 0)
                )
                item_score = (
                    (1 if item.name_hebrew else 0) +
                    (1 if item.name else 0) +
                    (1 if item.calories else 0) +
                    (1 if item.protein else 0) +
                    (1 if item.carbs else 0) +
                    (1 if item.fat else 0)
                )
                
                # If new item is better, mark existing as duplicate, otherwise mark new as duplicate
                if item_score > existing_score:
                    duplicates_to_remove.append(existing.id)
                    seen[key] = item
                else:
                    duplicates_to_remove.append(item.id)
            else:
                seen[key] = item
        
        # Remove duplicates
        removed_count = 0
        for item_id in duplicates_to_remove:
            item = db.query(MealBank).filter(MealBank.id == item_id).first()
            if item:
                db.delete(item)
                removed_count += 1
        
        db.commit()
        
        return {
            "message": f"Cleanup completed: {removed_count} duplicate items removed",
            "removed_count": removed_count,
            "duplicates_found": len(duplicates_to_remove)
        }
    except Exception as e:
        logger.error(f"Error cleaning up duplicates: {str(e)}", exc_info=True)
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to cleanup duplicates: {str(e)}"
        )

@router.get("/meal-bank/export/excel")
def export_meal_bank_excel(
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export all meal bank items to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Get all meal bank items (including public ones)
        meal_bank_items = db.query(MealBank).order_by(MealBank.macro_type, MealBank.name).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Meal Bank"
        
        # Header row
        headers = ["ID", "Name", "Name (Hebrew)", "Macro Type", "Calories (per 100g)", 
                   "Protein (g per 100g)", "Carbs (g per 100g)", "Fat (g per 100g)", 
                   "Created By", "Is Public", "Created At"]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for item in meal_bank_items:
            ws.append([
                item.id,
                item.name or "",
                item.name_hebrew or "",
                item.macro_type.value if item.macro_type else "",
                item.calories if item.calories is not None else "",
                item.protein if item.protein is not None else "",
                item.carbs if item.carbs is not None else "",
                item.fat if item.fat is not None else "",
                item.created_by,
                "Yes" if item.is_public else "No",
                item.created_at.strftime("%Y-%m-%d %H:%M:%S") if item.created_at else ""
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"meal_bank_export_{timestamp}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel export requires openpyxl library. Please install it."
        )
    except Exception as e:
        logger.error(f"Error exporting meal bank to Excel: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export meal bank: {str(e)}"
        )

@router.post("/meal-bank/import/excel")
async def import_meal_bank_excel(
    file: UploadFile = File(...),
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Parse Excel and return a preview of all rows with match status. No import until user reviews and confirms via /process."""
    if current_user.role != UserRole.TRAINER and current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only trainers can import meal bank items"
        )
    
    try:
        from openpyxl import load_workbook
        import unicodedata
        import re
        
        def normalize_hebrew(text: str) -> str:
            if not text:
                return ""
            text = unicodedata.normalize('NFKD', text)
            text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
            text = re.sub(r'\s+', ' ', text.strip().lower())
            return text
        
        def similarity_hebrew(text1: str, text2: str) -> float:
            """Return similarity 0..1. Exact match after norm = 1.0."""
            if not text1 or not text2:
                return 0.0
            norm1 = normalize_hebrew(text1)
            norm2 = normalize_hebrew(text2)
            if not norm1 or not norm2:
                return 0.0
            if norm1 == norm2:
                return 1.0
            # Levenshtein
            def lev(s1: str, s2: str) -> int:
                if len(s1) < len(s2):
                    return lev(s2, s1)
                if len(s2) == 0:
                    return len(s1)
                prev = list(range(len(s2) + 1))
                for i, c1 in enumerate(s1):
                    curr = [i + 1]
                    for j, c2 in enumerate(s2):
                        curr.append(min(prev[j + 1] + 1, curr[j] + 1, prev[j] + (1 if c1 != c2 else 0)))
                    prev = curr
                return prev[-1]
            d = lev(norm1, norm2)
            max_len = max(len(norm1), len(norm2))
            return 1.0 - (d / max_len) if max_len > 0 else 0.0
        
        # Possible duplicate if similarity >= this (catch more candidates for manual review)
        POSSIBLE_DUPLICATE_THRESHOLD = 0.60
        
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        existing_items = db.query(MealBank).filter(
            (MealBank.created_by == current_user.id) | (MealBank.is_public == True)
        ).all()
        
        rows_preview: List[Dict[str, Any]] = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or (not row[1] and not row[2]):
                    continue
                name = str(row[1]).strip() if len(row) > 1 and row[1] else None
                name_hebrew = str(row[2]).strip() if len(row) > 2 and row[2] else None
                if not name and not name_hebrew:
                    continue
                
                macro_type_str = str(row[3]).strip().lower() if len(row) > 3 and row[3] else "protein"
                if macro_type_str not in ["protein", "carb", "fat"]:
                    macro_type_str = "protein"
                
                item_data = {
                    "row_index": row_idx,
                    "name": name or name_hebrew,
                    "name_hebrew": name_hebrew if name_hebrew else None,
                    "macro_type": macro_type_str,
                    "calories": int(row[4]) if len(row) > 4 and row[4] and str(row[4]).strip() else None,
                    "protein": float(row[5]) if len(row) > 5 and row[5] and str(row[5]).strip() else None,
                    "carbs": float(row[6]) if len(row) > 6 and row[6] and str(row[6]).strip() else None,
                    "fat": float(row[7]) if len(row) > 7 and row[7] and str(row[7]).strip() else None,
                }
                
                matches: List[Dict[str, Any]] = []
                norm_hebrew = normalize_hebrew(name_hebrew or "")
                norm_name = normalize_hebrew(name or "")
                
                for existing in existing_items:
                    ex_hebrew = normalize_hebrew(existing.name_hebrew or "")
                    ex_name = normalize_hebrew(existing.name or "")
                    sim = 0.0
                    if norm_hebrew and ex_hebrew:
                        sim = max(sim, similarity_hebrew(name_hebrew or "", existing.name_hebrew or ""))
                    if norm_name and ex_name:
                        sim = max(sim, similarity_hebrew(name or "", existing.name or ""))
                    if sim >= POSSIBLE_DUPLICATE_THRESHOLD:
                        matches.append({
                            "id": existing.id,
                            "name": existing.name,
                            "name_hebrew": existing.name_hebrew,
                            "macro_type": existing.macro_type.value,
                            "calories": existing.calories,
                            "protein": existing.protein,
                            "carbs": existing.carbs,
                            "fat": existing.fat,
                            "similarity": round(sim, 2),
                        })
                
                status_key = "possible_duplicate" if matches else "new"
                rows_preview.append({
                    "row_index": row_idx,
                    "data": item_data,
                    "status": status_key,
                    "matches": matches,
                })
            except Exception as e:
                logger.error(f"Error processing row {row_idx}: {str(e)}")
        
        possible = sum(1 for r in rows_preview if r["status"] == "possible_duplicate")
        return {
            "rows": rows_preview,
            "message": f"Review {len(rows_preview)} rows. {possible} may match existing items; confirm each action before importing.",
        }
        
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel import requires openpyxl library. Please install it."
        )
    except Exception as e:
        logger.error(f"Error importing meal bank from Excel: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to import meal bank: {str(e)}"
        )

@router.post("/meal-bank/import/excel/process")
async def process_meal_bank_import(
    import_data: Dict[str, Any],
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Apply import from preview: decisions = { row_index: "skip" | "add" | "replace:<existing_id>" }."""
    if current_user.role != UserRole.TRAINER and current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only trainers can import meal bank items"
        )
    
    try:
        from app.models.meal_system import MeasurementType
        
        rows = import_data.get("rows", [])
        decisions = import_data.get("decisions", {})  # { row_index: "skip" | "add" | "replace:<id>" }
        
        imported_count = 0
        replaced_count = 0
        skipped_count = 0
        errors = []
        
        row_by_index = {r["row_index"]: r for r in rows}
        
        for row_index_str, decision in decisions.items():
            try:
                row_index = int(row_index_str)
                row_info = row_by_index.get(row_index)
                if not row_info:
                    skipped_count += 1
                    continue
                data = row_info.get("data", {})
                matches = row_info.get("matches", [])
                
                if decision == "skip":
                    skipped_count += 1
                    continue
                
                if decision == "add":
                    meal_bank_item = MealBank(
                        name=data.get("name"),
                        name_hebrew=data.get("name_hebrew"),
                        macro_type=MacroType(data.get("macro_type", "protein")),
                        calories=data.get("calories"),
                        protein=data.get("protein"),
                        carbs=data.get("carbs"),
                        fat=data.get("fat"),
                        measurement_type=MeasurementType.PER_100G,
                        serving_size=None,
                        created_by=current_user.id,
                        is_public=False
                    )
                    db.add(meal_bank_item)
                    imported_count += 1
                    continue
                
                if decision.startswith("replace:"):
                    try:
                        existing_id = int(decision.split(":", 1)[1])
                    except (ValueError, IndexError):
                        skipped_count += 1
                        continue
                    existing_item = db.query(MealBank).filter(MealBank.id == existing_id).first()
                    if not existing_item or (existing_item.created_by != current_user.id and current_user.role != UserRole.ADMIN):
                        skipped_count += 1
                        continue
                    existing_item.name = data.get("name")
                    existing_item.name_hebrew = data.get("name_hebrew")
                    existing_item.macro_type = MacroType(data.get("macro_type", "protein"))
                    existing_item.calories = data.get("calories")
                    existing_item.protein = data.get("protein")
                    existing_item.carbs = data.get("carbs")
                    existing_item.fat = data.get("fat")
                    replaced_count += 1
                else:
                    skipped_count += 1
            except Exception as e:
                errors.append(f"Row {row_index_str}: {str(e)}")
                skipped_count += 1
                logger.error(f"Error processing row {row_index_str}: {str(e)}")
        
        db.commit()
        
        return {
            "message": f"Import completed: {imported_count} items imported, {replaced_count} replaced, {skipped_count} skipped",
            "imported_count": imported_count,
            "replaced_count": replaced_count,
            "skipped_count": skipped_count,
            "errors": errors[:10]
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"Error processing meal bank import: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to process import: {str(e)}"
        )

@router.post("/plans/import/excel")
async def import_meal_plan_excel(
    file: UploadFile = File(...),
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Import meal plan from Excel file with duplicate detection"""
    if current_user.role != UserRole.TRAINER and current_user.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only trainers can import meal plans"
        )
    
    try:
        from openpyxl import load_workbook
        import unicodedata
        import re
        
        # Hebrew text normalization function
        def normalize_hebrew(text: str) -> str:
            """Normalize Hebrew text for matching (remove diacritics, handle variations)"""
            if not text:
                return ""
            # Remove diacritics (nikud)
            text = unicodedata.normalize('NFKD', text)
            text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
            # Remove extra spaces and convert to lowercase
            text = re.sub(r'\s+', ' ', text.strip().lower())
            return text
        
        # Fuzzy matching function
        def fuzzy_match_hebrew(text1: str, text2: str, threshold: float = 0.75) -> bool:
            """Check if two Hebrew texts are similar using Levenshtein distance"""
            if not text1 or not text2:
                return False
                
            norm1 = normalize_hebrew(text1)
            norm2 = normalize_hebrew(text2)
            
            if not norm1 or not norm2:
                return False
            
            # Exact match after normalization
            if norm1 == norm2:
                return True
            
            # Simple Levenshtein distance calculation
            def levenshtein(s1: str, s2: str) -> int:
                if len(s1) < len(s2):
                    return levenshtein(s2, s1)
                if len(s2) == 0:
                    return len(s1)
                
                previous_row = list(range(len(s2) + 1))
                for i, c1 in enumerate(s1):
                    current_row = [i + 1]
                    for j, c2 in enumerate(s2):
                        insertions = previous_row[j + 1] + 1
                        deletions = current_row[j] + 1
                        substitutions = previous_row[j] + (c1 != c2)
                        current_row.append(min(insertions, deletions, substitutions))
                    previous_row = current_row
                return previous_row[-1]
            
            distance = levenshtein(norm1, norm2)
            max_len = max(len(norm1), len(norm2))
            similarity = 1 - (distance / max_len) if max_len > 0 else 0
            
            # For short strings (3 chars or less), require exact match
            if max_len <= 3:
                return similarity == 1.0
            
            return similarity >= threshold
        
        # Read file content
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        
        # Sheet 1: Meal Plan Info
        if "Meal Plan Info" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="Excel file must contain 'Meal Plan Info' sheet")
        
        plan_sheet = wb["Meal Plan Info"]
        plan_row = list(plan_sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        
        client_id = int(plan_row[0]) if plan_row[0] else None
        plan_name = str(plan_row[1]).strip() if plan_row[1] else "Imported Meal Plan"
        plan_description = str(plan_row[2]).strip() if plan_row[2] else None
        number_of_meals = int(plan_row[3]) if plan_row[3] else 3
        total_calories = int(plan_row[4]) if plan_row[4] else None
        protein_target = int(plan_row[5]) if plan_row[5] else None
        carb_target = int(plan_row[6]) if plan_row[6] else None
        fat_target = int(plan_row[7]) if plan_row[7] else None
        
        if not client_id:
            raise HTTPException(status_code=400, detail="Client ID is required in Meal Plan Info sheet")
        
        # Verify client belongs to trainer
        from app.models.user import User
        client = db.query(User).filter(User.id == client_id).first()
        if not client:
            raise HTTPException(status_code=404, detail="Client not found")
        if current_user.role == UserRole.TRAINER and client.trainer_id != current_user.id:
            raise HTTPException(status_code=403, detail="You can only create meal plans for your clients")
        
        # Sheet 2: Meal Slots
        if "Meal Slots" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="Excel file must contain 'Meal Slots' sheet")
        
        slots_sheet = wb["Meal Slots"]
        meal_slots_data = []
        for row in slots_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            meal_slots_data.append({
                "name": str(row[0]).strip(),
                "time_suggestion": str(row[1]).strip() if row[1] else None,
                "target_calories": int(row[2]) if row[2] else None,
                "target_protein": float(row[3]) if row[3] else None,
                "target_carbs": float(row[4]) if row[4] else None,
                "target_fat": float(row[5]) if row[5] else None,
            })
        
        # Sheet 3: Macro Categories
        if "Macro Categories" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="Excel file must contain 'Macro Categories' sheet")
        
        categories_sheet = wb["Macro Categories"]
        macro_categories_data = {}
        for row in categories_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            meal_slot_name = str(row[0]).strip()
            macro_type = str(row[1]).strip().lower()
            if macro_type not in ["protein", "carb", "fat"]:
                continue
            
            if meal_slot_name not in macro_categories_data:
                macro_categories_data[meal_slot_name] = {}
            
            macro_categories_data[meal_slot_name][macro_type] = {
                "quantity_instruction": str(row[2]).strip() if row[2] else None,
                "calorie_goal": int(row[3]) if row[3] else None,
            }
        
        # Sheet 4: Food Options
        if "Food Options" not in wb.sheetnames:
            raise HTTPException(status_code=400, detail="Excel file must contain 'Food Options' sheet")
        
        food_sheet = wb["Food Options"]
        food_options_data = {}
        duplicate_matches = []
        
        # Get existing meal bank items for duplicate detection
        existing_items = db.query(MealBank).filter(
            MealBank.created_by == current_user.id
        ).all()
        
        for row in food_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            
            meal_slot_name = str(row[0]).strip()
            macro_type = str(row[1]).strip().lower()
            food_name = str(row[2]).strip() if row[2] else None
            food_name_hebrew = str(row[3]).strip() if row[3] else None
            
            if not food_name and not food_name_hebrew:
                continue
            
            # Check for duplicates using Hebrew text matching
            potential_duplicates = []
            if food_name_hebrew:
                for existing in existing_items:
                    if existing.name_hebrew and fuzzy_match_hebrew(food_name_hebrew, existing.name_hebrew):
                        potential_duplicates.append({
                            "id": existing.id,
                            "name": existing.name,
                            "name_hebrew": existing.name_hebrew
                        })
            
            if potential_duplicates:
                duplicate_matches.append({
                    "new_item": {"name": food_name, "name_hebrew": food_name_hebrew},
                    "matches": potential_duplicates
                })
            
            key = f"{meal_slot_name}_{macro_type}"
            if key not in food_options_data:
                food_options_data[key] = []
            
            food_options_data[key].append({
                "name": food_name,
                "name_hebrew": food_name_hebrew,
                "calories": int(row[4]) if row[4] else None,
                "protein": float(row[5]) if row[5] else None,
                "carbs": float(row[6]) if row[6] else None,
                "fat": float(row[7]) if row[7] else None,
                "serving_size": str(row[8]).strip() if row[8] else "100g",
                "measurement_type": str(row[9]).strip().lower() if row[9] else "per_100g",
            })
        
        # If duplicates found, return them for user to decide
        if duplicate_matches:
            return {
                "duplicates_found": True,
                "duplicate_matches": duplicate_matches[:20],  # Limit to 20 for response size
                "message": f"Found {len(duplicate_matches)} potential duplicate food items. Please review and decide whether to skip, merge, or rename."
            }
        
        # Create meal plan
        from app.models.meal_system import MeasurementType
        
        meal_plan = NewMealPlan(
            client_id=client_id,
            trainer_id=current_user.id,
            name=plan_name,
            description=plan_description,
            number_of_meals=number_of_meals,
            total_calories=total_calories,
            protein_target=protein_target,
            carb_target=carb_target,
            fat_target=fat_target,
            is_active=True
        )
        db.add(meal_plan)
        db.flush()
        
        # Create meal slots
        for idx, slot_data in enumerate(meal_slots_data):
            meal_slot = MealSlot(
                meal_plan_id=meal_plan.id,
                name=slot_data["name"],
                order_index=idx,
                time_suggestion=slot_data["time_suggestion"],
                target_calories=slot_data["target_calories"],
                target_protein=slot_data["target_protein"],
                target_carbs=slot_data["target_carbs"],
                target_fat=slot_data["target_fat"],
            )
            db.add(meal_slot)
            db.flush()
            
            # Create macro categories for this slot
            slot_categories = macro_categories_data.get(slot_data["name"], {})
            for macro_type in ["protein", "carb", "fat"]:
                category_data = slot_categories.get(macro_type, {})
                macro_category = MacroCategory(
                    meal_slot_id=meal_slot.id,
                    macro_type=MacroType(macro_type),
                    quantity_instruction=category_data.get("quantity_instruction"),
                    calorie_goal=category_data.get("calorie_goal"),
                )
                db.add(macro_category)
                db.flush()
                
                # Add food options for this category
                food_key = f"{slot_data['name']}_{macro_type}"
                food_items = food_options_data.get(food_key, [])
                
                for food_data in food_items:
                    food_option = FoodOption(
                        macro_category_id=macro_category.id,
                        name=food_data["name"],
                        name_hebrew=food_data["name_hebrew"],
                        calories=food_data["calories"],
                        protein=food_data["protein"],
                        carbs=food_data["carbs"],
                        fat=food_data["fat"],
                        serving_size=food_data["serving_size"],
                        measurement_type=MeasurementType.PER_100G if food_data["measurement_type"] == "per_100g" else MeasurementType.PER_PORTION,
                    )
                    db.add(food_option)
        
        db.commit()
        db.refresh(meal_plan)
        
        return {
            "message": "Meal plan imported successfully",
            "meal_plan_id": meal_plan.id,
            "meal_slots_count": len(meal_slots_data),
            "food_options_count": sum(len(items) for items in food_options_data.values())
        }
        
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel import requires openpyxl library. Please install it."
        )
    except Exception as e:
        logger.error(f"Error importing meal plan from Excel: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to import meal plan: {str(e)}"
        )
