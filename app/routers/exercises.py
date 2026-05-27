from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Form, Request
from fastapi import Request as FastAPIRequest
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional, Dict, Any
import logging
import json
import io
from datetime import datetime

from app.database import get_db
from app.services.workout_service import WorkoutService
from app.services.file_service import FileService
from app.auth.utils import get_current_user
from app.schemas.auth import UserResponse, UserRole
from app.schemas.workout import (
    ExerciseCreate, ExerciseUpdate, ExerciseResponse, ExerciseFilter
)
from app.models.workout import MuscleGroup, Exercise

logger = logging.getLogger(__name__)
router = APIRouter()

def get_file_service():
    """Dependency to get file service instance."""
    return FileService()

@router.post("/", response_model=ExerciseResponse, status_code=status.HTTP_201_CREATED)
async def create_exercise(
    request: FastAPIRequest,
    # Form fields for multipart/form-data
    exercise_json: Optional[str] = Form(None),
    image: Optional[UploadFile] = File(None),
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db),
    file_service: FileService = Depends(get_file_service)
):
    """
    Create a new exercise in the trainer's exercise bank.
    Supports both JSON (application/json) and multipart/form-data with optional image upload.
    
    For multipart/form-data:
    - Send exercise_json as a JSON string containing all exercise fields
    - Optionally send image file for exercise demonstration
    
    For application/json:
    - Send exercise data as JSON body
    """
    if current_user.role != UserRole.TRAINER:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only trainers can create exercises"
        )
    
    exercise_data: Optional[ExerciseCreate] = None
    
    # Check content type to determine how to parse
    content_type = request.headers.get("content-type", "")
    
    if "multipart/form-data" in content_type:
        # Handle multipart/form-data (for image uploads)
        if exercise_json:
            try:
                exercise_dict = json.loads(exercise_json)
                exercise_data = ExerciseCreate(**exercise_dict)
            except (json.JSONDecodeError, TypeError, ValueError) as e:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Invalid JSON in exercise_json field: {str(e)}"
                )
    elif "application/json" in content_type:
        # Handle JSON body
        try:
            body = await request.json()
            exercise_data = ExerciseCreate(**body)
        except (TypeError, ValueError) as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid exercise data: {str(e)}"
            )
    
    # Validate that we have exercise data
    if not exercise_data:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Exercise data is required. Send exercise_json as form field (for multipart/form-data) or JSON body (for application/json)."
        )
    
    # Handle image upload if provided
    image_path = None
    if image and image.filename:
        try:
            # Validate file
            is_valid, error_msg = await file_service.validate_file(image, file_type="image")
            if not is_valid:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=error_msg
                )
            
            # Save file - use temp directory first, then we'll move it after exercise creation
            file_result = await file_service.save_file(
                file=image,
                category="exercise_image",
                entity_id=0,  # Temporary ID
                process_image=True
            )
            image_path = file_result["original_path"]
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Error uploading exercise image: {str(e)}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Error uploading image: {str(e)}"
            )
    
    # Update exercise_data with image_path if uploaded
    exercise_dict = exercise_data.model_dump()
    if image_path:
        exercise_dict["image_path"] = image_path
    
    # Don't include image_path if video_url is already set (video takes priority)
    if exercise_dict.get("video_url") and image_path:
        # Still save image but video will be used for display
        pass
    
    exercise_data = ExerciseCreate(**exercise_dict)
    
    workout_service = WorkoutService(db)
    created_exercise = workout_service.create_exercise(exercise_data, current_user.id)
    
    # If image was uploaded and exercise was created successfully, we could rename the file
    # For now, the path is stored and will work
    if image_path and created_exercise.id:
        logger.info(f"Exercise {created_exercise.id} created with image: {image_path}")
    
    return created_exercise

@router.get("/", response_model=List[ExerciseResponse])
def get_exercises(
    trainer_id: Optional[int] = Query(None, description="Filter by trainer ID"),
    muscle_group: Optional[MuscleGroup] = Query(None, description="Filter by muscle group"),
    search: Optional[str] = Query(None, description="Search in exercise name, description, or instructions"),
    page: int = Query(1, ge=1, description="Page number"),
    size: int = Query(1000, ge=1, le=10000, description="Page size"),
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get exercises with filtering and pagination. All trainers can see all exercises."""
    workout_service = WorkoutService(db)
    
    # For trainers, don't filter by trainer_id unless explicitly requested
    # This ensures all trainers have access to all exercises
    filter_params = ExerciseFilter(
        trainer_id=trainer_id,  # Only filter if explicitly provided
        muscle_group=muscle_group,
        search=search,
        page=page,
        size=size
    )
    
    exercises, total = workout_service.get_exercises(filter_params)
    return exercises

@router.get("/{exercise_id}", response_model=ExerciseResponse)
def get_exercise(
    exercise_id: int,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get a specific exercise by ID."""
    workout_service = WorkoutService(db)
    exercise = workout_service.get_exercise(exercise_id)
    
    if not exercise:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Exercise not found"
        )
    
    return exercise

@router.put("/{exercise_id}", response_model=ExerciseResponse)
async def update_exercise(
    exercise_id: int,
    request: Request,
    # Form fields for multipart/form-data (when image is uploaded)
    exercise_json: Optional[str] = Form(None),
    image: Optional[UploadFile] = File(None),
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db),
    file_service: FileService = Depends(get_file_service)
):
    """
    Update an exercise (only by the trainer who created it).
    Supports both JSON (application/json) and multipart/form-data with optional image upload.
    
    For multipart/form-data:
    - Send exercise_json as a JSON string containing exercise fields to update
    - Optionally send image file to update/replace exercise image
    
    For application/json:
    - Send ExerciseUpdate as JSON body
    """
    if current_user.role != UserRole.TRAINER:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only trainers can update exercises"
        )
    
    exercise_data: Optional[ExerciseUpdate] = None
    
    # Handle multipart/form-data (for image uploads)
    if exercise_json is not None:
        # This is a multipart/form-data request
        try:
            exercise_dict = json.loads(exercise_json)
            exercise_data = ExerciseUpdate(**exercise_dict)
        except (json.JSONDecodeError, TypeError, ValueError) as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid JSON in exercise_json field: {str(e)}"
            )
    elif image is not None and image.filename:
        # Image uploaded but no exercise_json - need exercise_json
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="exercise_json field is required when uploading an image"
        )
    else:
        # Handle JSON body (backward compatibility - no form data)
        try:
            body = await request.json()
            exercise_data = ExerciseUpdate(**body)
        except (json.JSONDecodeError, ValueError, TypeError) as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid JSON body: {str(e)}"
            )
    
    # Validate we have exercise data
    if not exercise_data:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Exercise data is required. Send exercise_json as form field (for multipart/form-data) or JSON body (for application/json)."
        )
    
    # Handle image upload if provided
    image_path = None
    if image and image.filename:
        try:
            # Validate file
            is_valid, error_msg = await file_service.validate_file(image, file_type="image")
            if not is_valid:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=error_msg
                )
            
            # Save file
            file_result = await file_service.save_file(
                file=image,
                category="exercise_image",
                entity_id=exercise_id,
                process_image=True
            )
            image_path = file_result["original_path"]
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Error uploading exercise image: {str(e)}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Error uploading image: {str(e)}"
            )
    
    # Update exercise_data with image_path if uploaded
    exercise_dict = exercise_data.model_dump(exclude_unset=True)
    if image_path:
        exercise_dict["image_path"] = image_path
    
    # Don't include image_path if video_url is being set (video takes priority)
    if exercise_dict.get("video_url") and image_path:
        # Still save image but video will be used for display
        pass
    
    exercise_data = ExerciseUpdate(**exercise_dict)
    
    workout_service = WorkoutService(db)
    
    # Log for debugging
    logger.info(f"Updating exercise {exercise_id} for trainer {current_user.id} (role: {current_user.role})")
    
    # Check if exercise exists and who created it
    from app.models.workout import Exercise
    existing_exercise = db.query(Exercise).filter(Exercise.id == exercise_id).first()
    if existing_exercise:
        logger.info(f"Exercise {exercise_id} exists, created by user {existing_exercise.created_by}, current user is {current_user.id}")
    else:
        logger.warning(f"Exercise {exercise_id} not found in database")
    
    exercise = workout_service.update_exercise(exercise_id, exercise_data, current_user.id)
    
    if not exercise:
        logger.error(f"Failed to update exercise {exercise_id} for trainer {current_user.id} - exercise not found")
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Exercise not found"
        )
    
    if image_path and exercise.id:
        logger.info(f"Exercise {exercise.id} updated with image: {image_path}")
    
    return exercise

@router.delete("/{exercise_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_exercise(
    exercise_id: int,
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete an exercise (only by the trainer who created it)."""
    if current_user.role != UserRole.TRAINER:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only trainers can delete exercises"
        )
    
    # Check if exercise exists and belongs to the trainer
    from app.models.workout import Exercise
    from sqlalchemy import and_
    exercise = db.query(Exercise).filter(
        and_(
            Exercise.id == exercise_id,
            Exercise.created_by == current_user.id
        )
    ).first()
    
    if not exercise:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Exercise not found or you don't have permission to delete it"
        )
    
    # Check if exercise is used in any workout plans
    from app.models.workout import WorkoutExercise
    from app.models.workout_system import WorkoutExerciseV2, ExercisePersonalRecord
    
    workout_exercise_count = db.query(WorkoutExercise).filter(
        WorkoutExercise.exercise_id == exercise_id
    ).count()
    
    workout_exercise_v2_count = db.query(WorkoutExerciseV2).filter(
        WorkoutExerciseV2.exercise_id == exercise_id
    ).count()
    
    exercise_pr_count = db.query(ExercisePersonalRecord).filter(
        ExercisePersonalRecord.exercise_id == exercise_id
    ).count()
    
    if workout_exercise_count > 0 or workout_exercise_v2_count > 0 or exercise_pr_count > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Cannot delete exercise: It is currently used in {workout_exercise_count + workout_exercise_v2_count} workout plan(s) and {exercise_pr_count} personal record(s). Please remove it from all workout plans first."
        )
    
    # Try to delete the exercise directly
    try:
        db.delete(exercise)
        db.commit()
        return None  # 204 No Content
    except Exception as e:
        db.rollback()
        import traceback
        error_detail = str(e)
        logger.error(f"Error deleting exercise {exercise_id}: {error_detail}\n{traceback.format_exc()}")
        
        # Check if it's a foreign key constraint error
        if "foreign key" in str(e).lower() or "constraint" in str(e).lower():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Cannot delete exercise: It is still referenced in the database. Please ensure it's not used in any workout plans."
            )
        
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete exercise: {error_detail}"
        )

@router.get("/test")
async def test_exercises():
    return {"message": "Exercises router working"}

@router.get("/export/excel")
def export_exercises_excel(
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export all exercises to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from sqlalchemy.orm import joinedload
        
        # Get all exercises with muscle group relationship loaded
        exercises = db.query(Exercise).options(
            joinedload(Exercise.muscle_group_rel)
        ).order_by(Exercise.name).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Exercises"
        
        # Header row
        headers = ["ID", "Name", "Description", "Muscle Group", "Equipment Needed", 
                   "Instructions", "Category", "Video URL", "Image Path", "Created By", "Created At"]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for exercise in exercises:
            # Determine muscle group: use dynamic muscle group name if available, otherwise use string field
            muscle_group_name = ""
            if exercise.muscle_group_rel:
                # Exercise uses a dynamic muscle group
                muscle_group_name = exercise.muscle_group_rel.name
            elif exercise.muscle_group:
                # Exercise uses the string field (legacy enum or custom string)
                muscle_group_name = exercise.muscle_group
            
            ws.append([
                exercise.id,
                exercise.name or "",
                exercise.description or "",
                muscle_group_name,
                exercise.equipment_needed or "",
                exercise.instructions or "",
                exercise.category or "",
                exercise.video_url or "",
                exercise.image_path or "",
                exercise.created_by,
                exercise.created_at.strftime("%Y-%m-%d %H:%M:%S") if exercise.created_at else ""
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"exercises_export_{timestamp}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel export requires openpyxl library. Please install it."
        )
    except Exception as e:
        logger.error(f"Error exporting exercises to Excel: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export exercises: {str(e)}"
        )

def _normalize_hebrew(text: str) -> str:
    """Normalize Hebrew text for matching (remove nikud, lowercase, collapse spaces)."""
    import unicodedata
    import re
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = re.sub(r"\s+", " ", text.strip().lower())
    return text


def _similarity_hebrew(text1: str, text2: str) -> float:
    """Return similarity 0..1. Hebrew-safe; exact match after norm = 1.0."""
    import unicodedata
    import re
    if not text1 or not text2:
        return 0.0
    n1 = _normalize_hebrew(text1)
    n2 = _normalize_hebrew(text2)
    if not n1 or not n2:
        return 0.0
    if n1 == n2:
        return 1.0

    def lev(s1: str, s2: str) -> int:
        if len(s1) < len(s2):
            return lev(s2, s1)
        if len(s2) == 0:
            return len(s1)
        prev = list(range(len(s2) + 1))
        for i, c1 in enumerate(s1):
            curr = [i + 1]
            for j, c2 in enumerate(s2):
                curr.append(min(prev[j + 1] + 1, curr[j] + 1, prev[j] + (1 if c1 != c2 else 0)))
            prev = curr
        return prev[-1]

    d = lev(n1, n2)
    max_len = max(len(n1), len(n2))
    return 1.0 - (d / max_len) if max_len > 0 else 0.0


@router.post("/import/excel")
async def import_exercises_excel(
    file: UploadFile = File(...),
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Parse Excel and return a preview of all rows with match status (Hebrew-aware). No import until user reviews via /import/excel/process."""
    if current_user.role != UserRole.TRAINER:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only trainers can import exercises"
        )
    try:
        from openpyxl import load_workbook

        POSSIBLE_DUPLICATE_THRESHOLD = 0.60

        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active

        existing = db.query(Exercise).filter(Exercise.created_by == current_user.id).all()
        rows_preview: List[Dict[str, Any]] = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or len(row) < 2:
                    continue
                name = str(row[1]).strip() if row[1] else None
                if not name:
                    continue

                description = str(row[2]).strip() if len(row) > 2 and row[2] else None
                muscle_group = str(row[3]).strip() if len(row) > 3 and row[3] else "other"
                if muscle_group not in ("chest", "back", "shoulders", "biceps", "triceps", "legs", "core", "cardio", "full_body", "other"):
                    muscle_group = "other"
                equipment_needed = str(row[4]).strip() if len(row) > 4 and row[4] else None
                instructions = str(row[5]).strip() if len(row) > 5 and row[5] else None
                category = str(row[6]).strip() if len(row) > 6 and row[6] else None
                video_url = str(row[7]).strip() if len(row) > 7 and row[7] else None
                image_path = str(row[8]).strip() if len(row) > 8 and row[8] else None

                data = {
                    "row_index": row_idx,
                    "name": name,
                    "description": description,
                    "muscle_group": muscle_group,
                    "equipment_needed": equipment_needed,
                    "instructions": instructions,
                    "category": category,
                    "video_url": video_url,
                    "image_path": image_path,
                }
                matches: List[Dict[str, Any]] = []
                for ex in existing:
                    sim = _similarity_hebrew(name, ex.name or "")
                    if sim >= POSSIBLE_DUPLICATE_THRESHOLD:
                        matches.append({
                            "id": ex.id,
                            "name": ex.name,
                            "description": ex.description,
                            "muscle_group": ex.muscle_group,
                            "equipment_needed": ex.equipment_needed,
                            "similarity": round(sim, 2),
                        })
                status_key = "possible_duplicate" if matches else "new"
                rows_preview.append({
                    "row_index": row_idx,
                    "data": data,
                    "status": status_key,
                    "matches": matches,
                })
            except Exception as e:
                logger.error(f"Error processing exercise row {row_idx}: {str(e)}")

        possible = sum(1 for r in rows_preview if r["status"] == "possible_duplicate")
        return {
            "rows": rows_preview,
            "message": f"Review {len(rows_preview)} rows. {possible} may match existing exercises; confirm each action before importing.",
        }
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel import requires openpyxl library. Please install it."
        )
    except Exception as e:
        logger.error(f"Error importing exercises from Excel: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to import exercises: {str(e)}"
        )


@router.post("/import/excel/process")
async def process_exercises_import(
    import_data: Dict[str, Any],
    current_user: UserResponse = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Apply import from preview: decisions = { row_index: \"skip\" | \"add\" | \"replace:<existing_id>\" }."""
    if current_user.role != UserRole.TRAINER:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only trainers can import exercises"
        )
    try:
        rows = import_data.get("rows", [])
        decisions = import_data.get("decisions", {})
        row_by_index = {r["row_index"]: r for r in rows}
        workout_service = WorkoutService(db)
        imported_count = 0
        replaced_count = 0
        skipped_count = 0
        errors = []

        for row_index_str, decision in decisions.items():
            try:
                row_index = int(row_index_str)
                row_info = row_by_index.get(row_index)
                if not row_info:
                    skipped_count += 1
                    continue
                data = row_info.get("data", {})
                if decision == "skip":
                    skipped_count += 1
                    continue
                if decision == "add":
                    exercise_data = ExerciseCreate(
                        name=data.get("name", ""),
                        description=data.get("description"),
                        muscle_group=data.get("muscle_group", "other"),
                        equipment_needed=data.get("equipment_needed"),
                        instructions=data.get("instructions"),
                        category=data.get("category"),
                        video_url=data.get("video_url"),
                        image_path=data.get("image_path"),
                    )
                    workout_service.create_exercise(exercise_data, current_user.id)
                    imported_count += 1
                    continue
                if decision.startswith("replace:"):
                    try:
                        existing_id = int(decision.split(":", 1)[1])
                    except (ValueError, IndexError):
                        skipped_count += 1
                        continue
                    existing_item = db.query(Exercise).filter(Exercise.id == existing_id, Exercise.created_by == current_user.id).first()
                    if not existing_item:
                        skipped_count += 1
                        continue
                    exercise_update = ExerciseUpdate(
                        name=data.get("name", existing_item.name),
                        description=data.get("description"),
                        muscle_group=data.get("muscle_group", existing_item.muscle_group),
                        equipment_needed=data.get("equipment_needed"),
                        instructions=data.get("instructions"),
                        category=data.get("category"),
                        video_url=data.get("video_url"),
                        image_path=data.get("image_path"),
                    )
                    for field, value in exercise_update.model_dump(exclude_unset=True).items():
                        setattr(existing_item, field, value)
                    existing_item.created_by = current_user.id
                    replaced_count += 1
                else:
                    skipped_count += 1
            except Exception as e:
                errors.append(f"Row {row_index_str}: {str(e)}")
                skipped_count += 1
                logger.error(f"Error processing exercise row {row_index_str}: {str(e)}")

        db.commit()
        return {
            "message": f"Import completed: {imported_count} exercises created, {replaced_count} replaced, {skipped_count} skipped",
            "imported_count": imported_count,
            "replaced_count": replaced_count,
            "skipped_count": skipped_count,
            "errors": errors[:10],
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Error processing exercises import: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to process import: {str(e)}"
        ) 